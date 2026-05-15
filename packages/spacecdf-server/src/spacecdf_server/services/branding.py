"""SpaceCDF — Document branding for University of Ottawa / SEDTI.

Provides shared branding configuration and DOCX helper functions
for consistent headers, footers, and colour schemes across all exports.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any


@dataclass
class BrandingConfig:
    """Branding configuration applied to all exported documents."""
    university: str = "University of Ottawa"
    department: str = "SEDTI — Space Exploration & Design Technology Initiative"
    tool_name: str = "SpaceCDF"
    tool_version: str = "2.0"

    # Colours (UOttawa garnet + dark grey)
    primary_color: str = "#8B0000"    # UOttawa garnet
    secondary_color: str = "#2D2D2D"  # Dark grey
    accent_color: str = "#C8102E"     # Bright red accent
    header_bg: str = "#8B0000"
    header_text: str = "#FFFFFF"

    # Footer text
    footer_left: str = "University of Ottawa — SEDTI"
    footer_right: str = "SpaceCDF v2.0 — Concurrent Design Facility"

    # Classification
    classification: str = "UNCLASSIFIED — FOR ACADEMIC USE"

    # Logo paths (relative to project root)
    logo_path: str | None = None  # Set if logo file exists


# Singleton instance
BRANDING = BrandingConfig()


def get_branding() -> BrandingConfig:
    """Return the current branding configuration."""
    return BRANDING


def apply_docx_branding(doc: Any) -> None:
    """Apply the uOttawa SpaceCDF Facilitator's Book theme to a Document."""
    try:
        from spacecdf_agents.exporters.docs import theme as _theme
    except ImportError:
        return
    _theme.apply_styles(doc)
    _theme._set_page_geometry(doc)
    _theme.add_page_furniture(
        doc,
        running_title=f"{BRANDING.university} · {BRANDING.tool_name}",
        document_code="",
        footer_left=BRANDING.footer_left,
        footer_right="uOttawa SEDTI",
    )


def create_branded_docx(title: str, subtitle: str = "") -> Any:
    """Create a Document with the uOttawa course-style cover + page furniture."""
    try:
        from spacecdf_agents.exporters.docs import theme as _theme
    except ImportError:
        return None
    from datetime import datetime
    doc = _theme.new_document()
    _theme.add_cover_page(
        doc,
        title=title,
        subtitle=subtitle or BRANDING.tool_name,
        document_code="",
        study_name=BRANDING.tool_name,
        issue="1.0",
        date=datetime.now().strftime("%Y-%m-%d"),
        classification=BRANDING.classification,
        cohort=BRANDING.tool_name,
        publisher=f"{BRANDING.university} · {BRANDING.department}",
    )
    _theme.add_page_furniture(
        doc,
        running_title=f"{BRANDING.tool_name} — {title}",
        footer_left=BRANDING.footer_left,
        footer_right="uOttawa SEDTI",
    )
    # AIG (Peters 2023) acknowledgement on every branded document
    _theme.add_aig_acknowledgement(doc)
    return doc


def create_branded_xlsx(title: str) -> Any:
    """Create a new XLSX workbook with UOttawa/SEDTI branding.

    Returns an openpyxl Workbook with a branded title sheet.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Cover"

    b = BRANDING
    garnet_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    white_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")

    ws.merge_cells("A1:F1")
    ws["A1"].value = b.university
    ws["A1"].font = white_font
    ws["A1"].fill = garnet_fill
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    ws["A2"].value = b.department
    ws["A2"].font = Font(name="Calibri", size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A4:F4")
    ws["A4"].value = title
    ws["A4"].font = Font(name="Calibri", size=18, bold=True)
    ws["A4"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A6:F6")
    ws["A6"].value = b.classification
    ws["A6"].font = Font(name="Calibri", size=9, color="CC0000")
    ws["A6"].alignment = Alignment(horizontal="center")

    return wb
