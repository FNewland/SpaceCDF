"""SpaceCDF — XLSX Master Budget Generator.

Produces master_budget.xlsx with sheets for mass, power, data, cost WBS,
and compliance matrix. Uses openpyxl formulas (SUM, SUBTOTAL) so downstream
edits in Excel recompute totals automatically.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
TOTAL_FONT = Font(bold=True)
TOTAL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
STATUS_FILLS = {
    "compliant": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "green": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "marginal": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "amber": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "non_compliant": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "red": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "exceeded": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}


def _write_headers(ws, headers: list[str]) -> None:
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20


def _sheet_mass(wb: Workbook, budget: dict) -> None:
    ws = wb.create_sheet("Mass")
    _write_headers(
        ws,
        ["Subsystem", "Equipment", "Nominal_kg", "Margin_Pct", "With_Margin_kg"],
    )
    lines = budget.get("lines", []) if isinstance(budget, dict) else []
    for i, line in enumerate(lines, start=2):
        ws.cell(row=i, column=1, value=line.get("subsystem", ""))
        ws.cell(row=i, column=2, value=line.get("equipment", ""))
        ws.cell(row=i, column=3, value=float(line.get("nominal_value", 0) or 0))
        ws.cell(row=i, column=4, value=float(line.get("margin_percent", 0) or 0))
        ws.cell(row=i, column=5, value=f"=C{i}*(1+D{i}/100)")
    # Totals
    last = len(lines) + 1
    total_row = last + 1
    ws.cell(row=total_row, column=1, value="TOTAL").font = TOTAL_FONT
    if lines:
        ws.cell(row=total_row, column=3, value=f"=SUM(C2:C{last})").font = TOTAL_FONT
        ws.cell(row=total_row, column=5, value=f"=SUM(E2:E{last})").font = TOTAL_FONT
    ws.cell(row=total_row + 2, column=1, value="Allocation (kg)").font = TOTAL_FONT
    ws.cell(row=total_row + 2, column=3, value=float(budget.get("allocation", 0) or 0))
    ws.cell(row=total_row + 3, column=1, value="Margin %").font = TOTAL_FONT
    ws.cell(row=total_row + 3, column=3, value=float(budget.get("margin_percent", 0) or 0))


def _sheet_power(wb: Workbook, budget: dict) -> None:
    ws = wb.create_sheet("Power")
    _write_headers(
        ws,
        [
            "Subsystem", "Equipment", "Nominal_W", "Margin_Pct", "With_Margin_W",
            "Sunlight_W", "Eclipse_W",
        ],
    )
    lines = budget.get("lines", []) if isinstance(budget, dict) else []
    for i, line in enumerate(lines, start=2):
        nom = float(line.get("nominal_value", 0) or 0)
        ws.cell(row=i, column=1, value=line.get("subsystem", ""))
        ws.cell(row=i, column=2, value=line.get("equipment", ""))
        ws.cell(row=i, column=3, value=nom)
        ws.cell(row=i, column=4, value=float(line.get("margin_percent", 0) or 0))
        ws.cell(row=i, column=5, value=f"=C{i}*(1+D{i}/100)")
        ws.cell(row=i, column=6, value=nom)       # sunlight (assumed nominal)
        ws.cell(row=i, column=7, value=nom * 0.6)  # eclipse (heuristic 60%)
    last = len(lines) + 1
    total_row = last + 1
    ws.cell(row=total_row, column=1, value="TOTAL").font = TOTAL_FONT
    if lines:
        ws.cell(row=total_row, column=3, value=f"=SUM(C2:C{last})").font = TOTAL_FONT
        ws.cell(row=total_row, column=5, value=f"=SUM(E2:E{last})").font = TOTAL_FONT
        ws.cell(row=total_row, column=6, value=f"=SUM(F2:F{last})").font = TOTAL_FONT
        ws.cell(row=total_row, column=7, value=f"=SUM(G2:G{last})").font = TOTAL_FONT


def _sheet_data(wb: Workbook, budget: dict) -> None:
    ws = wb.create_sheet("Data")
    _write_headers(ws, ["Subsystem", "Source", "Generated_GB", "Stored_GB", "Downlinked_GB"])
    lines = budget.get("lines", []) if isinstance(budget, dict) else []
    for i, line in enumerate(lines, start=2):
        nom = float(line.get("nominal_value", 0) or 0)
        ws.cell(row=i, column=1, value=line.get("subsystem", ""))
        ws.cell(row=i, column=2, value=line.get("equipment", ""))
        ws.cell(row=i, column=3, value=nom)
        ws.cell(row=i, column=4, value=nom * 1.2)
        ws.cell(row=i, column=5, value=nom)
    last = len(lines) + 1
    total_row = last + 1
    if lines:
        ws.cell(row=total_row, column=1, value="TOTAL").font = TOTAL_FONT
        ws.cell(row=total_row, column=3, value=f"=SUM(C2:C{last})").font = TOTAL_FONT
        ws.cell(row=total_row, column=4, value=f"=SUM(D2:D{last})").font = TOTAL_FONT
        ws.cell(row=total_row, column=5, value=f"=SUM(E2:E{last})").font = TOTAL_FONT


def _sheet_cost(wb: Workbook, cost: dict) -> None:
    ws = wb.create_sheet("Cost WBS")
    _write_headers(
        ws,
        ["WBS_ID", "Name", "DDTE_kEUR", "Recurring_kEUR", "Total_kEUR"],
    )
    wbs = cost.get("wbs", []) if isinstance(cost, dict) else []
    for i, w in enumerate(wbs, start=2):
        ws.cell(row=i, column=1, value=w.get("wbs_id", ""))
        ws.cell(row=i, column=2, value=w.get("name", ""))
        ws.cell(row=i, column=3, value=float(w.get("ddte_keur", 0) or 0))
        ws.cell(row=i, column=4, value=float(w.get("recurring_keur", 0) or 0))
        ws.cell(row=i, column=5, value=f"=C{i}+D{i}")
    last = len(wbs) + 1
    total_row = last + 1
    if wbs:
        ws.cell(row=total_row, column=1, value="TOTAL").font = TOTAL_FONT
        ws.cell(row=total_row, column=3, value=f"=SUBTOTAL(9,C2:C{last})").font = TOTAL_FONT
        ws.cell(row=total_row, column=4, value=f"=SUBTOTAL(9,D2:D{last})").font = TOTAL_FONT
        ws.cell(row=total_row, column=5, value=f"=SUBTOTAL(9,E2:E{last})").font = TOTAL_FONT

    # Risk summary
    r = total_row + 2
    ws.cell(row=r, column=1, value="Risk Percentiles (kEUR)").font = TOTAL_FONT
    for offset, (label, key) in enumerate(
        [("P50", "p50_keur"), ("P70", "p70_keur"),
         ("P80", "p80_keur"), ("P90", "p90_keur")]
    ):
        ws.cell(row=r + 1 + offset, column=1, value=label)
        ws.cell(row=r + 1 + offset, column=2, value=float(cost.get(key, 0) or 0))


def _sheet_compliance(wb: Workbook, compliance: dict) -> None:
    ws = wb.create_sheet("Compliance")
    _write_headers(
        ws,
        ["Requirement_ID", "Description", "Threshold", "Achieved", "Margin_Pct", "Status"],
    )
    verifications = compliance.get("verifications", []) if isinstance(compliance, dict) else []
    requirements = {r.get("id"): r for r in (compliance.get("requirements", []) or [])}
    for i, v in enumerate(verifications, start=2):
        rid = v.get("requirement_id", "")
        ws.cell(row=i, column=1, value=rid)
        ws.cell(row=i, column=2, value=v.get("requirement_text", ""))
        req = requirements.get(rid, {})
        threshold = req.get("threshold") if req else None
        ws.cell(row=i, column=3, value=threshold if threshold is not None else "")
        av = v.get("achieved_value")
        ws.cell(row=i, column=4, value=av if isinstance(av, (int, float)) else "")
        mp = v.get("margin_percent")
        ws.cell(row=i, column=5, value=mp if isinstance(mp, (int, float)) else "")
        status = str(v.get("status", "")).lower()
        c = ws.cell(row=i, column=6, value=status.upper())
        if status in STATUS_FILLS:
            c.fill = STATUS_FILLS[status]


def generate_xlsx(context: dict[str, Any], output_path: Path) -> None:
    """Generate the master_budget.xlsx workbook."""
    wb = Workbook()
    # Remove the default sheet
    if wb.active is not None:
        wb.remove(wb.active)

    budgets = context.get("budgets", {}) or {}
    _sheet_mass(wb, budgets.get("mass", {}) or {})
    _sheet_power(wb, budgets.get("power", {}) or {})
    _sheet_data(wb, budgets.get("data", {}) or {})
    _sheet_cost(wb, context.get("cost", {}) or {})
    _sheet_compliance(wb, context.get("compliance", {}) or {})

    wb.save(str(output_path))
